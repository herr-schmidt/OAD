from numpy import *
from decimal import *
import openpyxl
import cplex
import openpyxl


def optimize(optimization_input_filepath, optimization_output_filepath):

    wb = openpyxl.load_workbook(optimization_input_filepath)
    sheet = wb['travel']
    x = []
    for row in sheet.iter_rows():
        x.append(row[0].value)

    y = []
    y = filter(None, x)
    y = [int(i) for i in y]
    ttt = []
    ttt.append([0]+y)

    pat = size(ttt)-1

    S = size(ttt)
    S2 = S

    wb = openpyxl.load_workbook(optimization_input_filepath)
    sheet = wb['visit']
    x = []
    for row in sheet.iter_rows():
        x.append(row[0].value)
    y = []
    y = filter(None, x)
    y = [int(i) for i in y]
    r1 = []
    r1.append([0]+y)

    # Service Time
    wb = openpyxl.load_workbook(optimization_input_filepath)
    sheet = wb.get_sheet_by_name(u'service')
    x = []
    for row in sheet.iter_rows():
        x.append(row[0].value)
    y = []
    y = filter(None, x)
    y = [int(i) for i in y]

    st = []
    for l in range(0, pat):
        st.append(y[l])
    st = [0]+st

    # Capacity
    wb = openpyxl.load_workbook(optimization_input_filepath)
    sheet = wb.get_sheet_by_name(u'capacity')
    x = []
    for row in sheet.iter_rows():
        x.append(row[0].value)
    y = []
    y = filter(None, x)
    y = [int(i) for i in y]

    a = []
    for l in range(0, len(y)):
        a.append(y[l])

    NO = len(y)

    # Skill
    wb = openpyxl.load_workbook(optimization_input_filepath)
    sheet = wb.get_sheet_by_name(u'SkillMatch')
    x = []
    for row in sheet.iter_rows():
        x.append(row[0].value)
    y = []
    y = filter(None, x)
    y = [int(i) for i in y]

    skl = []
    for l in range(0, len(y)):
        skl.append(y[l])

    MID = len(skl)

    # TeamSkill
    wb = openpyxl.load_workbook(optimization_input_filepath)
    sheet = wb.get_sheet_by_name(u'TeamsSkills')
    x = []
    for row in sheet.iter_rows():
        x.append(row[0].value)
    y = []
    y = filter(None, x)
    y = [int(i) for i in y]

    sk2 = []
    for l in range(0, len(y)):
        sk2.append(y[l])

    MS = max(sk2)

    # Patterns

    wb = openpyxl.load_workbook(optimization_input_filepath)
    sheet = wb.get_sheet_by_name(u'pattern')

    def iter_rows(sheet):
        for row in sheet.iter_rows():
            yield [cell.value for cell in row]
    aa = []
    aa = list(iter_rows(sheet))

    for i in range(0, int(size(aa, axis=0))):
        for j in range(0, int(size(aa, axis=1))):
            aa[i][j] = int(aa[i][j])

    A = array(aa)
    P = size(A, 0)
    S = size(ttt)
    S2 = S
    _list = {}
    ut = []
    Q = NO+1

    # DAys
    wb = openpyxl.load_workbook(optimization_input_filepath)
    sheet = wb.get_sheet_by_name(u'Days')
    x = []
    for row in sheet.iter_rows():
        x.append(row[0].value)
    y = []
    y = filter(None, x)
    y = [int(i) for i in y]

    DD = y[0]+1

    oo = 0
    zz = [0]

    demandd = [0]
    for i in range(1, S):
        demandd.append((ttt[0][i]+st[i]))

    for i in range(1, P):
        oo = 0
        for j in range(0, DD-1):
            b = aa[i]
            if b[j] == 2:
                b[j] = b[j]/2
            if b[j] == 3:
                b[j] = b[j]/3
            if b[j] == 4:
                b[j] = b[j]/4
            if b[j] == 5:
                b[j] = b[j]/5
            oo = oo+b[j]
        zz.append(oo)
        i = i+1

    # following part helps us to group of patterns that can be used for day d (has an entry greater than zer0)

    pat = []

    def ps(d):
        pat = [nonzero(A[:, d-1] == 1)]
        return pat[0][0]

    pat2 = []

    def ps2(d):
        pat2 = [nonzero(A[:, d-1] == 2)]
        return pat2[0][0]

    pat3 = []

    def ps3(d):
        pat3 = [nonzero(A[:, d-1] == 3)]
        return pat3[0][0]

    pat4 = []

    def ps4(d):
        pat4 = [nonzero(A[:, d-1] == 4)]
        return pat4[0][0]

    pat5 = []

    def ps5(d):
        pat5 = [nonzero(A[:, d-1] == 5)]
        return pat5[0][0]

    o = ()

    model = cplex.Cplex()
    model.parameters.mip.tolerances.absmipgap.set(float(0.01))
    # model.parameters.mip.limits.treememory.set(200)
    # model.parameters.timelimit.set(648000)

    for j in range(1, S2):
        for d in range(1, DD):
            for t in range(1, Q):
                model.variables.add(names=["Y"+str(j)+","+str(d)+","+str(t)])
                model.variables.set_types("Y"+str(j)+","+str(d)+","+str(t), "B")

    for j in range(1, S2):
        for p in range(1, P):
            model.variables.add(names=["Z"+str(j)+","+str(p)])
            model.variables.set_types("Z"+str(j)+","+str(p), "B")

    for j in range(1, S2):
        for t in range(1, Q):
            model.variables.add(names=["X"+str(j)+","+str(t)])
            model.variables.set_types("X"+str(j)+","+str(t), "B")

    # to be able to use rindex and avoid reversing
    sk2_as_string = "".join(map(str, sk2))

    for i in range(skl[0], S2):
        for t in range(1, sk2_as_string.rindex("1")+2):
            model.variables.advanced.tighten_upper_bounds("X"+str(i)+","+str(t), 0)

    for i in range(skl[1], S2):
        for t in range(1, sk2_as_string.rindex("2")+2):
            model.variables.advanced.tighten_upper_bounds("X"+str(i)+","+str(t), 0)

    for i in range(skl[2], S2):
        for t in range(1, sk2_as_string.rindex("3")+2):
            model.variables.advanced.tighten_upper_bounds("X"+str(i)+","+str(t), 0)

    for i in range(skl[3], S2):
        for t in range(1, sk2_as_string.rindex("4")+2):
            model.variables.advanced.tighten_upper_bounds("X"+str(i)+","+str(t), 0)

    for t in range(1, Q):
        for d in range(1, DD):
            model.variables.add(names=["W"+str(t)+","+str(d)])

    model.variables.add(names=["h"])

    model.objective.set_sense(model.objective.sense.minimize)
    model.objective.set_linear("h", 1)

    ##constraints##
    ##1##
    for j in range(1, S2):
        for d in range(1, DD):
            model.linear_constraints.add(names=["const6"+","+str(j)+","+str(d)])
            for t in range(1, Q):
                model.linear_constraints.set_coefficients("const6"+","+str(j)+","+str(d), "Y"+str(j)+","+str(d)+","+str(t), 1)
            if j in range(1, skl[0]):
                for p in ps(d):
                    model.linear_constraints.set_coefficients("const6"+","+str(j)+","+str(d), "Z"+str(j)+","+str(p), -1)
            if j in range(skl[0], skl[1]):
                for p in ps2(d):
                    model.linear_constraints.set_coefficients("const6"+","+str(j)+","+str(d), "Z"+str(j)+","+str(p), -1)
            if j in range(skl[1], skl[2]):
                for p in ps3(d):
                    model.linear_constraints.set_coefficients("const6"+","+str(j)+","+str(d), "Z"+str(j)+","+str(p), -1)
            if j in range(skl[2], skl[3]):
                for p in ps4(d):
                    model.linear_constraints.set_coefficients("const6"+","+str(j)+","+str(d), "Z"+str(j)+","+str(p), -1)
            if j in range(skl[3], S2):
                for p in ps5(d):
                    model.linear_constraints.set_coefficients("const6"+","+str(j)+","+str(d), "Z"+str(j)+","+str(p), -1)
            model.linear_constraints.set_rhs("const6"+","+str(j)+","+str(d), 0)
            model.linear_constraints.set_senses("const6"+","+str(j)+","+str(d), "E")

    ##3##
    ##assignment##
    for j in range(1, S2):
        model.linear_constraints.add(names=["const10"+str(j)])
        for t in range(1, Q):
            model.linear_constraints.set_coefficients("const10"+str(j), "X"+str(j)+","+str(t), 1)
        model.linear_constraints.set_rhs("const10"+str(j), 1)
        model.linear_constraints.set_senses("const10"+str(j), "E")
    ##4##
    ##PATTERN##
    for j in range(1, S2):
        model.linear_constraints.add(names=["const9"+","+str(j)])
        for p in range(1, P):
            model.linear_constraints.set_coefficients("const9"+","+str(j), "Z"+str(j)+","+str(p), 1)
        model.linear_constraints.set_rhs("const9"+","+str(j), 1)
        model.linear_constraints.set_senses("const9"+","+str(j), "E")

    ##5##
    for j in range(1, S2):
        for d in range(1, DD):
            for t in range(1, Q):
                model.linear_constraints.add(names=["const3"+","+str(j)+","+str(d)+","+str(t)])
                model.linear_constraints.set_coefficients("const3"+","+str(j)+","+str(d)+","+str(t), "Y"+str(j)+","+str(d)+","+str(t), 1)
                model.linear_constraints.set_coefficients("const3"+","+str(j)+","+str(d)+","+str(t), "X"+str(j)+","+str(t), -1)
                model.linear_constraints.set_rhs("const3"+","+str(j)+","+str(d)+","+str(t), 0)
                model.linear_constraints.set_senses("const3"+","+str(j)+","+str(d)+","+str(t), "L")

    ####6##
    for j in range(1, S2):
        model.linear_constraints.add(names=["const18"+","+str(j)])
        for t in range(1, Q):
            for d in range(1, DD):
                model.linear_constraints.set_coefficients("const18"+","+str(j), "Y"+str(j)+","+str(d)+","+str(t), 1)
        model.linear_constraints.set_rhs("const18"+","+str(j), r1[0][j])
        model.linear_constraints.set_senses("const18"+","+str(j), "E")

    ##7##
    ## workload ##
    for t in range(1, Q):
        for d in range(1, DD):
            model.linear_constraints.add(names=["const11"+str(t)+","+str(d)])
            model.linear_constraints.set_coefficients("const11"+str(t)+","+str(d), "W"+str(t)+","+str(d), -1)
            for j in range(1, S2):
                model.linear_constraints.set_coefficients("const11"+str(t)+","+str(d), "Y"+str(j)+","+str(d)+","+str(t), demandd[j])
            model.linear_constraints.set_senses("const11"+str(t)+","+str(d), "E")
    ##8##
    ## minimum utilization ##
    for t in range(1, Q):
        model.linear_constraints.add(names=["const12"+str(t)])
        for d in range(1, DD):
            model.linear_constraints.set_coefficients("const12"+str(t), "W"+str(t)+","+str(d), float(1./(a[t-1]*(DD-1))))
        model.linear_constraints.set_coefficients("const12"+str(t), "h", -1)
        model.linear_constraints.set_rhs("const12"+str(t), 0)
        model.linear_constraints.set_senses("const12"+str(t), "L")

    ##9##
    for j in range(1, S2):
        model.linear_constraints.add(names=["const8"+","+str(j)])
        for p in range(1, P):
            model.linear_constraints.set_coefficients("const8"+","+str(j), "Z"+str(j)+","+str(p), zz[p])
        model.linear_constraints.set_rhs("const8"+","+str(j), r1[0][j])
        model.linear_constraints.set_senses("const8"+","+str(j), "E")

    ##10##

    #### capacity ##
    for t in range(1, Q):
        for d in range(1, DD):
            model.linear_constraints.add(names=["const19"+str(t)+","+str(d)])
            model.linear_constraints.set_coefficients("const19"+str(t)+","+str(d), "W"+str(t)+","+str(d), 1)
            model.linear_constraints.set_rhs("const19"+str(t)+","+str(d), a[t-1])
            model.linear_constraints.set_senses("const19"+str(t)+","+str(d), "L")

    # lancement de la résolution
    model.solve()

    # affichage des résultats
    # print "MAX UTI:       ", model.solution.get_objective_value()

    # for j in range (1,S2):
    # for t in range(1,Q):
    # print(j,t)
    # print(model.solution.get_values("X"+str(j)+","+str(t)))

    # for j in range (1,S2):
    # for d in range(1,DD):
    # for t in range(1,Q):
    # if model.solution.get_values("Y"+str(j)+","+str(d)+","+str(t))>0:
    # print(j,t,d)
    ##
    # for t in range (1,Q):
    # for d in range(1,DD):
    # print(t,d)
    # print(model.solution.get_values("W"+str(t)+","+str(d)))

    # q={}
    # tt=0
    # for j in range (1,S2):
    # for d in range(1,DD):
    # for t in range(1,Q):
    # if model.solution.get_values("Y"+str(j)+","+str(d)+","+str(t))>0:
    # q[tt]=(j,t,d)
    # tt=tt+1

    ##print (q)

    # for j in range (1,S2):
    # for p in range(1,P):
    # if model.solution.get_values("Z"+str(j)+","+str(p))>0:
    # print(j,p)
    # print(model.solution.get_values("Z"+str(j)+","+str(p)))

    ptlst = []
    for j in range(1, S2):
        for p in range(1, P):
            if model.solution.get_values("Z"+str(j)+","+str(p)) > 0.1:
                # print(j,p)
                ptlst.append((j, p))
    # print(model.solution.get_values("Z"+str(j)+","+str(p)))

    c = []
    ##the following is used to write the patietn list of nurses####
    for j in range(1, Q):
        for i in range(1, S):
            if model.solution.get_values("X"+str(i)+","+str(j)) > 0.1:
                c.append(i)
        _list[("n"+str(j))] = (c)
        c = []

    id = {}
    for x in range(1, Q):
        id["nrs{0}".format(x)] = _list[("n"+str(x))]

    ln = {}
    for x in range(1, Q):
        ln["l{0}".format(x)] = len(_list[("n"+str(x))])

    Wop = zeros(shape=(NO, 5))

    for t in range(1, Q):
        for d in range(1, DD):
            Wop[t-1][d-1] = model.solution.get_values("W"+str(t)+","+str(d))

    import xlwt

    font0 = xlwt.Font()
    font0.name = 'Times New Roman'
    font0.colour_index = 2
    font0.bold = True

    wb = xlwt.Workbook()
    ws = wb.add_sheet('Pattern')

    for k in range(0, len(ptlst)):
        for j in range(0, 2):
            if j == 0:
                ws.write(k, 0, ptlst[k][j])
            else:
                ws.write(k, 1, ptlst[k][j])

    wb.save(optimization_output_filepath)

    ws = wb.add_sheet('Operator list')

    for x in range(1, Q):
        for k in range(0, len(_list[("n"+str(x))])):
            ws.write(x-1, k, id["nrs{0}".format(x)][k])

    ws = wb.add_sheet('Number of patients')
    for x in range(1, Q):
        ws.write(x-1, 0, len(_list[("n"+str(x))]))
    wb.save(optimization_output_filepath)

    ws = wb.add_sheet('Work load')

    for x in range(0, size(Wop, 0)):
        for y in range(0, size(Wop, 1)):
            ws.write(x, y, Wop[x][y])

    wb.save(optimization_output_filepath)
